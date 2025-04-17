import pandas as pd
from datetime import datetime, timedelta
from openpyxl import load_workbook
import os
from dotenv import load_dotenv

# Load environment variables for file paths
load_dotenv()
RECORD_DIR = os.getenv("RECORD_DIR", "path/to/record")
DONE_DIR = os.getenv("DONE_DIR", "path/to/done")

# Calculate date range for last Monday to Friday
today = datetime.now()
last_monday = today - timedelta(days=6)  # Last Monday (6 days before Sunday)
last_friday = last_monday + timedelta(days=4)  # Last Friday
start_str = last_monday.strftime("%d%b%Y")  # e.g., "24Mar2025"
end_str = last_friday.strftime("%d%b%Y")    # e.g., "28Mar2025"

# File configurations for processing
files_to_process = [
    {
        "base_name": "System1 CPU Utilization ",
        "target_file_path": os.path.join(RECORD_DIR, "System1 CPU Utilization.xlsx"),
        "start_row": 1,
        "start_col": 5  # Column E
    },
    {
        "base_name": "System1 Disk Utilization ",
        "target_file_path": os.path.join(RECORD_DIR, "System1 Disk Utilization.xlsx"),
        "start_row": 1,
        "start_col": 8  # Column H
    },
    {
        "base_name": "System2 CPU Utilization ",
        "target_file_path": os.path.join(RECORD_DIR, "System2 CPU Utilization.xlsx"),
        "start_row": 1,
        "start_col": 5  # Column E
    },
    {
        "base_name": "System2 Disk Utilization ",
        "target_file_path": os.path.join(RECORD_DIR, "System2 Disk Utilization.xlsx"),
        "start_row": 1,
        "start_col": 8  # Column H
    }
]

def process_file(base_name, target_path, start_row, start_col):
    """Process a single file by extracting data from .xls and updating .xlsx."""
    source_file_name = f"{base_name}{start_str}-{end_str}.xls"
    source_file_path = os.path.join(RECORD_DIR, source_file_name)
    new_save_path = os.path.join(DONE_DIR, f"{base_name}{start_str}-{end_str}.xlsx")

    print(f"\nProcessing file: {source_file_name}")

    # Step 1: Read source .xls file
    try:
        dfs = pd.read_html(source_file_path)
        print(f"Successfully read source file: {source_file_name}")
        df = dfs[0]
        # Extract data from A1, remove empty rows/columns
        data = df.dropna(how='all').loc[:, df.notna().any()]
        print(f"Extracted data: {data.shape[0]} rows, {data.shape[1]} columns")
        print("Data preview:")
        print(data.head())
    except ValueError as e:
        print(f"Failed to read HTML: {e}")
        return False
    except FileNotFoundError:
        print(f"Source file not found: {source_file_path}")
        return False
    except Exception as e:
        print(f"Unexpected error: {e}")
        return False

    # Step 2: Update target .xlsx file
    try:
        wb = load_workbook(target_path)
        ws = wb.active
        print(f"Opened target file: {target_path}")
        print(f"Using worksheet: {ws.title}")

        # Paste data, handling merged cells
        for i in range(data.shape[0]):
            for j in range(data.shape[1]):
                value = data.iloc[i, j]
                if pd.notna(value):
                    # Convert string numbers to numeric types
                    try:
                        if isinstance(value, str) and value.replace('.', '', 1).lstrip('-').replace('e', '', 1).replace('+', '', 1).isdigit():
                            value = float(value)
                        elif isinstance(value, str) and value.lstrip('-').isdigit():
                            value = int(value)
                    except (ValueError, TypeError):
                        pass  # Keep as string if not numeric

                    target_row = start_row + i
                    target_col = start_col + j
                    cell = ws.cell(row=target_row, column=target_col)

                    # Check if cell is in a merged range
                    is_merged = False
                    merged_top_left = None
                    for merged_range in ws.merged_cells.ranges:
                        min_row, min_col, max_row, max_col = merged_range.bounds
                        if (min_row <= target_row <= max_row and
                            min_col <= target_col <= max_col):
                            is_merged = True
                            merged_top_left = (min_row, min_col)
                            break

                    if is_merged and (target_row, target_col) == merged_top_left:
                        cell.value = value
                        print(f"Wrote to merged cell top-left at R{target_row}C{target_col} ({chr(64 + target_col)}{target_row}): {value}")
                    elif not is_merged:
                        cell.value = value
                        print(f"Wrote to cell at R{target_row}C{target_col} ({chr(64 + target_col)}{target_row}): {value}")
                    else:
                        print(f"Skipped merged cell at R{target_row}C{target_col} ({chr(64 + target_col)}{target_row}): {value}")

        # Step 3: Save as new .xlsx file
        wb.save(new_save_path)
        print(f"Saved new file: {new_save_path}")
        print(f"Original target file unchanged: {target_path}")
        return True
    except FileNotFoundError:
        print(f"Target file not found: {target_path}")
        return False
    except Exception as e:
        print(f"Error processing target file: {e}")
        return False

# Process all files
for file_config in files_to_process:
    success = process_file(
        base_name=file_config["base_name"],
        target_path=file_config["target_file_path"],
        start_row=file_config["start_row"],
        start_col=file_config["start_col"]
    )
    if not success:
        print(f"Failed to process {file_config['base_name']}, continuing to next file")

# Final output
print(f"\nLast Monday: {start_str}")
print(f"Last Friday: {end_str}")